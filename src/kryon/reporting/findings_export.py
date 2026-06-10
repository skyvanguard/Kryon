"""Tabular export of findings to CSV / Excel — for client tracking.

Banks/SMBs track remediation in spreadsheets, not PDFs. This normalizes both
compliance ``CheckResult`` rows and offensive ``engage.Finding`` rows into one
flat schema and writes CSV (stdlib, always available) or XLSX (needs the
optional ``openpyxl`` from the ``reporting`` extra, with severity colouring and
a summary sheet).
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from kryon.reporting.export import _REPORTS_DIR, _safe_slug

# Flat, client-facing column order. Stable — clients build trackers on it.
COLUMNS: tuple[str, ...] = (
    "id",
    "title",
    "severity",
    "framework",
    "control",
    "host",
    "status",
    "remediation",
    "evidence",
)

# Severity → Excel fill colour (hex, no '#'). Order also drives the summary sheet.
_SEVERITY_ORDER: tuple[str, ...] = ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO")
_SEVERITY_FILL: dict[str, str] = {
    "CRITICAL": "C00000",
    "HIGH": "E26B0A",
    "MEDIUM": "BF9000",
    "LOW": "538135",
    "INFO": "808080",
}
_COL_WIDTH: dict[str, int] = {
    "id": 16,
    "title": 42,
    "severity": 12,
    "framework": 14,
    "control": 14,
    "host": 18,
    "status": 10,
    "remediation": 50,
    "evidence": 50,
}
_EVIDENCE_CAP = 500
_TITLE_CAP = 160


@dataclass(frozen=True)
class FindingRow:
    """One normalized, client-facing finding row."""

    id: str
    title: str
    severity: str
    framework: str
    control: str
    host: str
    status: str
    remediation: str
    evidence: str

    def as_dict(self) -> dict[str, str]:
        return {c: getattr(self, c) for c in COLUMNS}


def _s(obj: Any, attr: str, default: str = "") -> str:
    return str(getattr(obj, attr, default) or default)


def from_check_result(result: Any, framework: str = "") -> FindingRow:
    """Adapt a compliance ``CheckResult`` into a client row."""
    evidence = (_s(result, "evidence_stdout") or _s(result, "evidence_command"))[:_EVIDENCE_CAP]
    return FindingRow(
        id=_s(result, "control_id"),
        title=_s(result, "control_title")[:_TITLE_CAP],
        severity=_s(result, "severity").upper(),
        framework=framework,
        control=_s(result, "control_id"),
        host=_s(result, "host"),
        status=_s(result, "verdict").upper(),
        remediation=_s(result, "remediation_static"),
        evidence=evidence,
    )


def from_engage_finding(finding: Any) -> FindingRow:
    """Adapt an offensive ``engage.Finding`` into a client row."""
    return FindingRow(
        id=_s(finding, "rule_id"),
        title=_s(finding, "message")[:_TITLE_CAP],
        severity=_s(finding, "severity").upper(),
        framework="",
        control=_s(finding, "cwe"),
        host=_s(finding, "host") or _s(finding, "target_host"),
        status="OPEN",
        remediation=_s(finding, "remediation"),
        evidence=_s(finding, "evidence")[:_EVIDENCE_CAP],
    )


def from_intel_finding(finding: Any) -> FindingRow:
    """Adapt an ``intelligence.models.Finding`` (used by ``kryon report``) into
    a client row. That model uses id/title/description/affected_asset/cve."""
    return FindingRow(
        id=_s(finding, "id"),
        title=_s(finding, "title")[:_TITLE_CAP],
        severity=_s(finding, "severity").upper(),
        framework=_s(finding, "tool_source"),
        control=_s(finding, "cve") or _s(finding, "mitre"),
        host=_s(finding, "affected_asset"),
        status=(_s(finding, "validation_status") or "OPEN").upper(),
        remediation=_s(finding, "remediation"),
        evidence=_s(finding, "evidence")[:_EVIDENCE_CAP],
    )


def _write_csv(rows: list[FindingRow], path: Path) -> None:
    # utf-8-sig so Excel renders accents (PCI/Spanish reports) correctly.
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow(row.as_dict())


def _write_xlsx(rows: list[FindingRow], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - exercised only without the extra
        raise RuntimeError(
            'Excel export needs openpyxl. Install it with `pip install "kryon[reporting]"` '
            "or `pip install openpyxl`, or export with --format csv."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws.append([c.upper() for c in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")

    sev_idx = COLUMNS.index("severity") + 1
    for row in rows:
        data = row.as_dict()
        ws.append([data[c] for c in COLUMNS])
        fill = _SEVERITY_FILL.get(row.severity.upper())
        if fill:
            cell = ws.cell(row=ws.max_row, column=sev_idx)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color="FFFFFF", bold=True)

    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTH.get(col, 18)
    ws.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary.append(["Severity", "Count"])
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.severity.upper()] = counts.get(row.severity.upper(), 0) + 1
    for sev in _SEVERITY_ORDER:
        if sev in counts:
            summary.append([sev, counts[sev]])
    summary.append(["TOTAL", len(rows)])
    wb.save(path)


def export_findings(
    rows: list[FindingRow],
    fmt: str = "csv",
    client_name: str = "",
    report_type: str = "findings",
) -> Path:
    """Write rows to the reports dir as CSV or XLSX. Returns the path."""
    fmt = fmt.lower()
    if fmt not in ("csv", "xlsx"):
        raise ValueError(f"unsupported export format {fmt!r}; use 'csv' or 'xlsx'")
    rows = list(rows)
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    slug = _safe_slug(client_name) if client_name else "report"
    filename = f"{slug}_{_safe_slug(report_type, 'findings')}_{ts}.{fmt}"
    path = (_REPORTS_DIR / filename).resolve()
    if not str(path).startswith(str(_REPORTS_DIR.resolve())):
        raise ValueError("Invalid report path")
    if fmt == "xlsx":
        _write_xlsx(rows, path)
    else:
        _write_csv(rows, path)
    return path
